import openpyxl as xl  # type: ignore
from datetime import datetime
from openpyxl.chart import PieChart, Reference, BarChart  # type: ignore

def create_wb(filename='Expenses.xlsx'):
    try:
        workbook = xl.load_workbook(filename)
        if 'Expenses' not in workbook.sheetnames:
            sheet = workbook.create_sheet('Expenses')
            sheet['A1'] = 'Date'
            sheet['B1'] = 'Category'
            sheet['C1'] = 'Amount'
            sheet['D1'] = 'Total'
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 15  # Fixed typo here
            workbook.save(filename)
            print("Created a new sheet 'Expenses'.")
        else:
            print("Sheet 'Expenses' already exists.")
    except FileNotFoundError:
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = "Expenses"
        sheet['A1'] = 'Date'
        sheet['B1'] = 'Category'
        sheet['C1'] = 'Amount'
        sheet['D1'] = 'Total'
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        workbook.save(filename)
        print(f"Excel file '{filename}' created successfully with sheet 'Expenses'!")
    return workbook

def add_items(workbook, date, category, expense, filename='Expenses.xlsx'):
    sheet = workbook['Expenses']
    next_row = sheet.max_row + 1
    sheet[f"A{next_row}"] = date
    sheet[f"B{next_row}"] = category
    sheet[f"C{next_row}"] = expense
    
    # Calculate the new total expense
    total_expense = sum(sheet[f"C{i}"].value for i in range(2, next_row + 1))
    sheet[f"D{next_row}"] = total_expense  # Update total column
    
    # Sort the expenses by date in ascending order
    sort_expenses_by_date(workbook, filename)
    
    workbook.save(filename)
    print("Expense added successfully and sorted by date!")

def sort_expenses_by_date(workbook, filename='Expenses.xlsx'):
    sheet = workbook['Expenses']
    # Extract all data except header
    data = []
    for row in range(2, sheet.max_row + 1):
        date = sheet[f"A{row}"].value
        category = sheet[f"B{row}"].value
        amount = sheet[f"C{row}"].value
        data.append((date, category, amount))
    
    # Sort data by date
    data_sorted = sorted(data, key=lambda x: datetime.strptime(x[0], "%Y-%m-%d"))
    
    # Clear existing data
    for row in range(2, sheet.max_row + 1):
        sheet[f"A{row}"].value = None
        sheet[f"B{row}"].value = None
        sheet[f"C{row}"].value = None
        sheet[f"D{row}"].value = None
    
    # Re-insert sorted data and recalculate totals
    for idx, (date, category, amount) in enumerate(data_sorted, start=2):
        sheet[f"A{idx}"] = date
        sheet[f"B{idx}"] = category
        sheet[f"C{idx}"] = amount
        # Recalculate total up to this row
        total_expense = sum(row[2] for row in data_sorted[:idx-1])
        sheet[f"D{idx}"] = total_expense
    
    workbook.save(filename)

def validate_date(input_date):
    try:
        datetime.strptime(input_date, "%Y-%m-%d")
        return True
    except ValueError:
        print("Incorrect date format. Please use YYYY-MM-DD.")
        return False

def validate_category(input_category, valid_categories):
    if input_category in valid_categories:
        return True
    else:
        print("Invalid category. Please choose from the list.")
        return False

def validate_expense(input_expense):
    try:
        expense = float(input_expense)
        if expense <= 0:
            print("Expense amount must be positive.")
            return False
        return True
    except ValueError:
        print("Invalid expense amount. Please enter a number.")
        return False

def remove_expense(workbook, row_number, filename='Expenses.xlsx'):
    sheet = workbook['Expenses']
    if 1 < row_number <= sheet.max_row:
        sheet.delete_rows(row_number)
        # Recalculate totals after deletion
        recalculate_totals(workbook, filename)
        workbook.save(filename)
        print(f"Row {row_number} deleted successfully and totals recalculated.")
    else:
        print("Invalid row number. Please enter a valid row.")

def recalculate_totals(workbook, filename='Expenses.xlsx'):
    sheet = workbook['Expenses']
    total = 0
    for row in range(2, sheet.max_row + 1):
        amount = sheet[f"C{row}"].value
        if amount:
            total += amount
            sheet[f"D{row}"] = total
    workbook.save(filename)

def summary(workbook, filename='Expenses.xlsx'):
    sheet = workbook['Expenses']
    summary_dict = {}
    for row in range(2, sheet.max_row + 1):
        category = sheet[f"B{row}"].value
        amount = sheet[f"C{row}"].value
        if category in summary_dict:
            summary_dict[category] += amount
        else:
            summary_dict[category] = amount

    if 'Summary' in workbook.sheetnames:
        summary_sheet = workbook['Summary']
        workbook.remove(summary_sheet)
    summary_sheet = workbook.create_sheet('Summary')
    summary_sheet['A1'] = 'Category'
    summary_sheet['B1'] = 'Total Amount'

    for idx, (category, total) in enumerate(summary_dict.items(), start=2):
        summary_sheet[f'A{idx}'] = category
        summary_sheet[f'B{idx}'] = total

    pie = PieChart()
    labels = Reference(summary_sheet, min_col=1, min_row=2, max_row=summary_sheet.max_row)
    data = Reference(summary_sheet, min_col=2, min_row=1, max_row=summary_sheet.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Expense Distribution by Category"
    summary_sheet.add_chart(pie, "D2")

    bar = BarChart()
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = "Total Expenses per Category"
    bar.y_axis.title = 'Amount'
    bar.x_axis.title = 'Category'
    summary_sheet.add_chart(bar, "D20")

    workbook.save(filename)
    print("Summary with charts generated successfully!")

def add_category(valid_categories, category):
    if category not in valid_categories:
        valid_categories.append(category)
        print(f"Category '{category}' added successfully.")
    else:
        print(f"Category '{category}' already exists.")

def modify_category(valid_categories, old_category, new_category, workbook, filename='Expenses.xlsx'):
    if old_category in valid_categories:
        valid_categories[valid_categories.index(old_category)] = new_category
        update_category_in_expenses(workbook, old_category, new_category, filename)
        print(f"Category '{old_category}' modified to '{new_category}'.")
    else:
        print(f"Category '{old_category}' does not exist.")

def update_category_in_expenses(workbook, old_category, new_category, filename='Expenses.xlsx'):
    sheet = workbook['Expenses']
    for row in range(2, sheet.max_row + 1):
        if sheet[f"B{row}"].value == old_category:
            sheet[f"B{row}"].value = new_category
    workbook.save(filename)
    summary(workbook, filename)  # Update summary after modification

def remove_expenses_by_category(workbook, category_to_remove, filename='Expenses.xlsx'):
    sheet = workbook['Expenses']
    rows_to_delete = []
    
    for row in range(2, sheet.max_row + 1):
        if sheet[f"B{row}"].value == category_to_remove:
            rows_to_delete.append(row)

    for row in reversed(rows_to_delete):
        sheet.delete_rows(row)
    
    recalculate_totals(workbook, filename)
    workbook.save(filename)
    print(f"All expenses under the category '{category_to_remove}' have been deleted.")

def remove_category(workbook, valid_categories, filename='Expenses.xlsx'):
    category_to_remove = input("Enter the category to remove: ")
    if category_to_remove in valid_categories:
        valid_categories.remove(category_to_remove)
        print(f"Category '{category_to_remove}' removed successfully.")
        remove_expenses_by_category(workbook, category_to_remove, filename)
        summary(workbook, filename)
    else:
        print(f"Category '{category_to_remove}' does not exist.")

def category_menu(valid_categories, workbook, filename='Expenses.xlsx'):
    while True:
        print("\n----Category Management----")
        print("1. Add Category")
        print("2. Modify Category")
        print("3. Remove Category")
        print("4. Back to Main Menu")
        cat_choice = input("Enter your choice: ")
        if cat_choice == '1':
            new_category = input("Enter the new category: ")
            add_category(valid_categories, new_category)
            summary(workbook, filename)  # Update summary after adding category
        elif cat_choice == '2':
            old_category = input("Enter the category to modify: ")
            new_category = input("Enter the new name for the category: ")
            modify_category(valid_categories, old_category, new_category, workbook, filename)
            summary(workbook, filename)  # Update summary after modifying category
        elif cat_choice == '3':
            remove_category(workbook, valid_categories, filename)
        elif cat_choice == '4':
            break
        else:
            print("Invalid choice. Please try again.")
            continue

def display_menu():
    print("\n........Expense Tracker Menu........")
    print("1. Add Expense")
    print("2. Delete Expense")
    print("3. Category Management")
    print("4. View Summary Report")
    print("5. Exit")

def add_expenses_menu(workbook, valid_categories, filename='Expenses.xlsx'):
    while True:
        date = input("Enter the Date (YYYY-MM-DD): ")
        while not validate_date(date):
            date = input("Enter the Date (YYYY-MM-DD): ")
        print(f"Available Categories: {', '.join(valid_categories)}")
        category = input("Enter the expense category: ")
        while not validate_category(category, valid_categories):
            category = input("Enter the expense category: ")
        expense = input("Enter the amount spent: ")
        while not validate_expense(expense):
            expense = input("Enter the amount spent: ")
        add_items(workbook, date, category, float(expense), filename)
        cont = input("Do you want to add another expense? (yes/no): ").strip().lower()
        if cont != 'yes':
            print("Expenses Added! Check it out in the Excel sheet.")
            break

def main():
    filename = 'Expenses.xlsx'
    workbook = create_wb(filename)
    valid_categories = ["Food", "Transport", "Entertainment", "Groceries", "Others"]
    while True:
        display_menu()
        choice = input("Enter your choice (1-5): ").strip()
        if choice == '1':
            add_expenses_menu(workbook, valid_categories, filename)
        elif choice == '2':
            try:
                row_number = int(input("Enter the row number to be removed: "))
                remove_expense(workbook, row_number, filename)
                summary(workbook, filename)
            except ValueError:
                print("Invalid input. Please enter a numeric row number.")
        elif choice == '3':
            category_menu(valid_categories, workbook, filename)
        elif choice == '4':
            summary(workbook, filename)
        elif choice == '5':
            print("Thank you for using the Expense Tracker! Goodbye!")
            break
        else:
            print("Invalid choice. Please select a valid option from the menu.")

if __name__ == "__main__":
    main()
